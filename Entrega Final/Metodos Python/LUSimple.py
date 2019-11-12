from openpyxl import Workbook
from funciones import *
import numpy as np
Excel=Workbook()
LU_Simple=Excel.active
LU_Simple.title="LU_Simple"
def LUSimple(matriz,vectorSol):
    n=len(matriz)
    L=np.zeros((n,n))
    U=np.zeros((n,n))
    for i in range(n):
        L[i][i]=1
        U[0][i]=matriz[0][i]
        for j in range(i+1,n):
            mult=(matriz[j][i]) / (matriz[i][i])
            L[j][i]=mult
            for k in range(n):
                matriz[j][k]=matriz[j][k]-(mult*matriz[i][k])
                U[j][k]=matriz[j][k]
    vecz=sustProgre(L,vectorSol)
    sln=sustRegre(U,vecz) #Vector donde estan las x
    LU_Simple.cell(row=1,column=1,value="Matriz L final ")
    for i in range(n):
        for j in range(n):
            LU_Simple.cell(row=i+2,column=j+1,value=L[i][j])
    cont=n+3
    LU_Simple.cell(row=cont,column=1,value="Matriz U final ")
    for i in range(n):
        for j in range(n):
            LU_Simple.cell(row=cont+2,column=j+1,value=U[i][j])
        cont+=1
    cont+=3
    LU_Simple.cell(row=cont,column=1,value="La Solución al Sistema es: ")
    xi=1
    for i in range(n):
        LU_Simple.cell(row=cont+2,column=1,value=("x" + str(xi)))
        LU_Simple.cell(row=cont+2,column=2,value=sln[i])
        xi+=1
        cont+=1
    Excel.save(filename="LUSimple.xlsx")
    return(L,U,sln)
#n=len(L)
#print("Matriz L\n"+str(L))
#print("Matriz U\n"+str(U))
#print("La solución al sistema es: ")
#for i in range(n):
    #print("x"+str(i+1)+": "+str(sln[i]))