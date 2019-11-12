import numpy as np
from funciones import *
import sympy as sp
from openpyxl import Workbook
Excel=Workbook()
SplineCubico=Excel.active
SplineCubico.title="Spline_Cubico"
def splinecubico(xi,yi):
    x=sp.Symbol('x')
    n=4*(len(xi)-1)
    A=np.zeros((n,n))
    b=np.zeros(n)
    cont=2
    ini=0
    sucont=-1
    for i in range(0,n,4):
        for j in range(ini,cont):
            A[j][i]=(xi[j])**3
            A[j][i+1]=(xi[j])**2
            A[j][i+2]=(xi[j])
            A[j][i+3]=1
            b[j]=yi[j]
        if ini>=2:
            ini=cont
            sucont+=1
        if ini==0:
            ini+=2
        cont+=1
    c=1
    for i in range(((len(xi))),(cont+sucont)):
        A[i]=A[c]
        c+=1
    d=1
    ini2=((len(xi))+2)
    fin2=(cont+sucont+1)
    while fin2-ini2>1:
        ini2+=1
    for i in range(0,n-4,4):
        for j in range(ini2,fin2):
            A[j][i]=(3*((xi[d])**2))
            A[j][i+1]=((xi[d])*2)
            A[j][i+2]=1
            A[j][i+4]=-A[j][i]
            A[j][i+5]=-A[j][i+1]
            A[j][i+6]=-A[j][i+2]
            d+=1
            ini2+=1
            fin2+=1
    cont3=sucont
    ini3=0
    fin3=4
    for i in range((cont-1),(cont+cont3)):
        for j in range(ini3,fin3):
            if j==ini3:
                A[i][j+4]=-A[i][j]
            if j!=ini3:
                A[i][j+4]=(-(A[i][j]))
        ini3+=4
        fin3+=4
    d=1
    for i in range(0,n-4,4):
        for j in range(ini2,fin2):
            A[j][i]=((xi[d])*6)
            A[j][i+1]=2
            A[j][i+4]=-A[j][i]
            A[j][i+5]=-A[j][i+1]
            d+=1
            ini2+=1
            fin2+=1
    ini4=0
    fin4=ini4+1
    r=0
    for i in range(n-2,n):
        for j in range(ini4,fin4):
            A[i][j]=((xi[r])*6)
            A[i][j+1]=2
        b[i]=0
        ini4=n-4
        fin4=ini4+1
        nx=len(xi)
        r=nx-1
    #print(A)
    #print(b)
    GauSimple(A,b)
    sln=sustRegre(A,b)
    SplineCubico.cell(row=1,column=1,value="Coeficientes de los trazadores ")
    con=0
    sum=3
    for i in range(0,len(A)-1,4):
        SplineCubico.cell(row=sum,column=1,value=sln[i])
        SplineCubico.cell(row=sum,column=2,value=sln[i+1])
        SplineCubico.cell(row=sum,column=3,value=sln[i+2])
        SplineCubico.cell(row=sum,column=4,value=sln[i+3])
        SplineCubico.cell(row=sum,column=6,value="Intervalo ["+str(xi[con])+","+str(xi[con+1])+"]")
        #print("Intervalo ["+str(xi[con])+","+str(xi[con+1])+"]")
        #print((sln[i]*(x**3))+((sln[i+1])*(x**2))+((sln[i+2])*x)+sln[i+3])
        con+=1
        sum+=1
    sum+=2
    SplineCubico.cell(row=sum,column=1,value="Trazadores")
    sum+=2
    con=0
    for i in range(0,len(A)-1,4):
        SplineCubico.cell(row=sum,column=1,value=str((sln[i]*(x**3))+((sln[i+1])*(x**2))+((sln[i+2])*x)+sln[i+3]))
        SplineCubico.cell(row=sum,column=6,value="Intervalo ["+str(xi[con])+","+str(xi[con+1])+"]")
        sum+=1
        con+=1
    Excel.save(filename="Spline_Cubico.xlsx")
#splinecubico(xi,yi)llamar a la funcion