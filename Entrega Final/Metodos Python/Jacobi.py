import numpy as np
from openpyxl import Workbook
Excel=Workbook()
jacobi=Excel.active
jacobi.title="Jacobi"
def jacobi2(A,b,x0,tol,Nmax):
    n=len(A)
    D_inv,L,U=np.zeros((n,n)),np.zeros((n,n)),np.zeros((n,n))
    for i in range(n):
        D_inv[i][i]=1/(A[i][i])
        for j in range(i+1,n):
            L[j][i]=-A[j][i]
            U[i][j]=-A[i][j]
    Tj=D_inv@(L+U)
    Cj=D_inv@b
    vp,vep=np.linalg.eig(Tj)
    p=max(abs(vp))
    #print("T\n"+str(Tj))
    jacobi.cell(row=1,column=1,value="Tj")
    for i in range(n):
        for j in range(n):
            jacobi.cell(row=i+2,column=j+1,value=Tj[i][j])
    #print("C\n"+str(Cj))
    cont=n+3
    jacobi.cell(row=cont,column=1,value="Cj")
    for i in range(n):
        jacobi.cell(row=cont+2,column=i+1,value=Cj[i])
    #print("Radio espectral: "+str(p))
    jacobi.cell(row=2,column=n+2,value="Radio Espectral")
    jacobi.cell(row=2,column=n+3,value=p)
    cont+=4
    jacobi.cell(row=cont,column=1,value="Iter")
    jacobi.cell(row=cont,column=2,value="Error")
    for i in range(n):
        jacobi.cell(row=cont,column=3+i,value=("x"+str(i+1)))
    cont+=1
    xant=x0
    for k in range(Nmax):
        xactual=(Tj@xant)+Cj
        #print("Iteración "+str(k+1))
        jacobi.cell(row=cont,column=1,value=k+1)
        SumErr=0
        for i in range(n):
            #print("x"+str(i+1)+": "+str(xactual[i]))
            jacobi.cell(row=cont,column=i+3,value=xactual[i])
            SumErr+=((xactual[i]-xant[i])**2)
        xant=xactual
        Error=np.sqrt(SumErr)
        #print("Error: "+str(Error))
        jacobi.cell(row=cont,column=2,value=Error)
        cont+=1
        if Error < tol:
            break
    Excel.save(filename="Jacobi.xlsx")
#jacobi2(A,b,x0,tol,Nmax) llamar a la funcion