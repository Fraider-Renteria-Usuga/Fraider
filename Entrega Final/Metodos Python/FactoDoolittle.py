from openpyxl import Workbook
from funciones import *
import numpy as np
Excel=Workbook()
doolitle=Excel.active
doolitle.title="Doolitle"
def Doolittle(A,b):
    n=len(A)
    L=np.identity(n)
    U=np.zeros((n,n))
    for k in range(n):
        for i in range(n):
            SumaL=0
            Sumakk=0
            for p in range(k):
                SumaL+=L[i][p]*U[p][k]
                Sumakk+=L[k][p]*U[p][k]
            U[k][k]=A[k][k]-Sumakk
            L[i][k]=(A[i][k]-SumaL)/U[k][k]
        for j in range(k+1,n):
            SumaU=0
            for p in range(k):
                SumaU+=L[k][p]*U[p][j]
            U[k][j]=(A[k][j]-SumaU)/L[k][k]
    z=sustProgre(L,b)
    sln=sustRegre(U,z)
    doolitle.cell(row=1,column=1,value="Matriz L final ")
    for i in range(n):
        for j in range(n):
            doolitle.cell(row=i+2,column=j+1,value=L[i][j])
    cont=n+3
    doolitle.cell(row=cont,column=1,value="Matriz U final ")
    for i in range(n):
        for j in range(n):
            doolitle.cell(row=cont+2,column=j+1,value=U[i][j])
        cont+=1
    cont+=3
    doolitle.cell(row=cont,column=1,value="La Solución al Sistema es: ")
    xi=1
    for i in range(n):
        doolitle.cell(row=cont+2,column=1,value=("x"+str(xi)))
        doolitle.cell(row=cont+2,column=2,value=sln[i])
        xi+=1
        cont+=1
    Excel.save(filename="Doolitle.xlsx")
    return(L,U,sln)
#n=len(L)
#Comando para imprimir de una manera clara al usuario
#print("Matriz L\n"+str(L))
#print("Matriz U\n"+str(U))
#print("La solución al sistema es: ")
#for i in range(n):
    #print("x"+str(i+1)+": "+str(sln[i]))