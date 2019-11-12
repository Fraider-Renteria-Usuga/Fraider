from funciones import f
import numpy as np
from sympy import *
from openpyxl import Workbook
Excel=Workbook()
newton=Excel.active
newton.title="NewtonDif_Divididas"
def newtondif(xi,yi):
    n=len(xi)
    x=symbols('x')
    cont=0
    for k in range(n):
        cont+=1
        for i in range(cont,n):
            p=yi[i][cont-1]
            u=yi[i-1][cont-1]
            col=(p-u)/(xi[i]-xi[i-cont])
            yi[i].append(col)
    #print("xi   f[xi] ")
    newton.cell(row=1,column=1,value="Tabla de Diferencias divididas")
    newton.cell(row=3,column=1,value="n")
    newton.cell(row=3,column=2,value="Xi")
    newton.cell(row=3,column=3,value="f[xi]")
    p=0
    cont=4
    for i in range(n):
        #print(str(xi[i])+str(yi[i]))
        newton.cell(row=cont,column=1,value=i)
        newton.cell(row=cont,column=2,value=xi[i])
        for j in range(len(yi[i])):
            newton.cell(row=cont,column=j+3,value=yi[i][j])
        cont+=1
        t=yi[i][i]
        for k in range(i):
            t*=(x-xi[k])
        p+=t
    #print("El polinomio es: \n" + str(p))
    cont+=3
    newton.cell(row=cont,column=1,value="Polinomio ")
    newton.cell(row=cont+1,column=1,value=str(p))
    Excel.save(filename="NewtonDif_Divididas.xlsx")
#xi=[-1.,0.,3.,4.]
#yi=[[15.5],[3.],[8.],[1.]] Ejemplo de ejecucion
#newtondif(xi,yi)