import numpy as np
from openpyxl import Workbook
Excel=Workbook()
Sor=Excel.active
Sor.title="Sor"
def sor(A,b,x0,tol,Nmax,w):
    n=len(A)
    D,L,U=np.zeros((n,n)),np.zeros((n,n)),np.zeros((n,n))
    for i in range(n):
        D[i][i]=(A[i][i])
        for j in range(i+1,n):
            L[j][i]=-A[j][i]
            U[i][j]=-A[i][j]
    Tw=(np.linalg.inv((D-(w*L))))@(((1-w)*D)+(w*U))
    Cw=(w*(np.linalg.inv((D-(w*L)))))@b
    #print("Tw\n"+str(Tw))
    Sor.cell(row=1,column=1,value="Tw")
    for i in range(n):
        for j in range(n):
            Sor.cell(row=i+2,column=j+1,value=Tw[i][j])
    #print("Cw\n"+str(Cw))
    cont=n+3
    Sor.cell(row=cont,column=1,value="Cw")
    for i in range(n):
        Sor.cell(row=cont+2,column=i+1,value=Cw[i])
    vp,vep=np.linalg.eig(Tw)
    p=max(abs(vp))
    #print("\nRadio espectral: "+str(p))
    Sor.cell(row=2,column=n+2,value="Radio Espectral")
    Sor.cell(row=2,column=n+3,value=p)
    cont+=4
    Sor.cell(row=cont,column=1,value="Iter")
    Sor.cell(row=cont,column=2,value="Error")
    for i in range(n):
        Sor.cell(row=cont,column=3+i,value=("x" + str(i + 1)))
    cont+=1
    xant=x0
    for k in range(Nmax):
        xactual=(Tw@xant)+Cw
        #print("Iteración "+str(k+1))
        Sor.cell(row=cont,column=1,value=k+1)
        SumErr=0
        for i in range(n):
            #print("x"+str(i+1)+": "+str(xactual[i]))
            Sor.cell(row=cont,column=i+3,value=xactual[i])
            SumErr+=((xactual[i]-xant[i])**2)
        xant=xactual
        Error=np.sqrt(SumErr)
        #print("Error: "+str(Error))
        Sor.cell(row=cont,column=2,value=Error)
        cont+=1
        if Error < tol:
            break
    Excel.save(filename="SOR.xlsx")
#sor(A,b,x0,tol,Nmax,w) Llamar a la funcion