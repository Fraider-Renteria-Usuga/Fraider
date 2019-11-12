from openpyxl import Workbook
from funciones import *
import numpy as np
Excel=Workbook()
crout=Excel.active
crout.title="Crout"
def Crout(A,b):
  n=len(A)
  L=np.zeros((n,n))
  U=np.identity(n)
  for k in range(n):
    for i in range(n):
      SumaL=0
      SumaU=0
      for p in range(k):
        SumaL+=L[i][p]*U[p][k]
        SumaU+=L[k][p]*U[p][i]
      L[i][k]=A[i][k]-SumaL
      if L[k][k]!=0:
          U[k][i]=(A[k][i]-SumaU)/L[k][k]
  z=sustProgre(L,b)
  sln=sustRegre(U,z)
  crout.cell(row=1,column=1,value="Matriz L final ")
  for i in range(n):
    for j in range(n):
      crout.cell(row=i+2,column=j+1,value=L[i][j])
  cont=n+3
  crout.cell(row=cont,column=1,value="Matriz U final ")
  for i in range(n):
    for j in range(n):
      crout.cell(row=cont+2,column=j+1,value=U[i][j])
    cont+=1
  cont+=3
  crout.cell(row=cont,column=1,value="La Solución al Sistema es: ")
  xi=1
  for i in range(n):
    crout.cell(row=cont+2,column=1,value=("x"+str(xi)))
    crout.cell(row=cont+2,column=2,value=sln[i])
    xi+=1
    cont+=1
  Excel.save(filename="Crout.xlsx")
  return(L,U,sln)
#n=len(L)
#Comando para imprimir de una manera clara al usuario
#print("Matriz L\n"+str(L))
#print("Matriz U\n"+str(U))
#print("La solución al sistema es: ")
#for i in range(n):
    #print("x"+str(i+1)+": "+str(sln[i]))