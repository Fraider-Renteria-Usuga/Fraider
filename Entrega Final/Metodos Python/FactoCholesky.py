import numpy as np
from math import sqrt
from funciones import *
from openpyxl import Workbook
Excel=Workbook()
Cholesky=Excel.active
Cholesky.title="Cholesky"
def cholesky(A,b):
    n=len(A)
    L=np.zeros((n,n))
    U=np.zeros((n,n))
    for k in range(n):
        for i in range(n):
            SumaL=0
            Sumakk=0
            for p in range(k):
                SumaL+=L[i][p]*U[p][k]
                Sumakk+=L[k][p]*U[p][k]
            L[k][k]=sqrt(A[k][k]-Sumakk)
            U[k][k]=L[k][k]
            L[i][k]=(A[i][k]-SumaL)/U[k][k]
        for j in range(k+1,n):
            SumaU=0
            for p in range(k):
                SumaU+=L[k][p]*U[p][j]
            U[k][j]=(A[k][j]-SumaU)/L[k][k]
    z=sustProgre(L,b)
    sln=sustRegre(U,z)
    Cholesky.cell(row=1,column=1,value="Matriz L final ")
    for i in range(n):
        for j in range(n):
            Cholesky.cell(row=i+2,column=j+1,value=L[i][j])
    cont=n+3
    Cholesky.cell(row=cont,column=1,value="Matriz U final ")
    for i in range(n):
        for j in range(n):
            Cholesky.cell(row=cont+2,column=j+1,value=U[i][j])
        cont+=1
    cont+=3
    Cholesky.cell(row=cont,column=1,value="La Solución al Sistema es: ")
    xi=1
    for i in range(n):
        Cholesky.cell(row=cont+2,column=1,value=("x" + str(xi)))
        Cholesky.cell(row=cont+2,column=2,value=sln[i])
        xi+=1
        cont+=1
    Excel.save(filename="Cholesky.xlsx")
    return(L,U,sln)
#L,U,sln=cholesky(A,b) #  Llamar a la funcion
#n=len(A)
#Comando para imprimir de una manera clara al usuario
#print("Matriz L\n"+str(L))
#print("Matriz U\n"+str(U))
#for i in range(n):
    #print("x"+str(n+1)+": "+str(sln[i]))