import numpy as np
from funciones import *
import sympy as sp
from openpyxl import Workbook
Excel=Workbook()
SplineCuadratico=Excel.active
SplineCuadratico.title="Spline_Cuadratico"
def splinecuadratico(xi,yi):
    x=sp.Symbol('x')
    n=3*(len(xi)-1)
    A=np.zeros((n,n))
    b=np.zeros(n)
    cont=2
    ini=0
    sucont=-1
    for i in range(0,n,3):
        for j in range(ini,cont):
            A[j][i]=(xi[j])**2
            A[j][i+1]=xi[j]
            A[j][i+2]=1
            b[j]=yi[j]
        if ini>=2:
            ini=cont
            sucont+=1
        if ini==0:
            ini+=2
        cont+=1
    c=1
    for i in range(((len(xi))),(cont+sucont)):
        A[i]=A[c]
        c+=1
    d=1
    ini2=((len(xi))+2)
    fin2=(cont+sucont+1)
    while fin2-ini2>1:
        ini2+=1
    for i in range(0,n-3,3):
        for j in range(ini2,fin2):
            A[j][i]=((xi[d])*2)
            A[j][i+1]=1
            A[j][i+3]=-A[j][i]
            A[j][i+4]=-A[j][i+1]
            d+=1
            ini2+=1
            fin2+=1
    cont3=sucont
    ini3=0
    fin3=3
    for i in range((cont-1),(cont+cont3)):
        exp=1
        for j in range(ini3,fin3):
            if j==ini3:
                A[i][j+3]=-A[i][j]
            if j!=ini3:
                A[i][j+3]=(-((A[i][j])**(exp)))
                exp-=1
        ini3+=3
        fin3+=3
    A[n-1][0]=2
    #print(A)
    GauSimple(A,b)
    sln=sustRegre(A,b)
    SplineCuadratico.cell(row=1,column=1,value="Coeficientes de los trazadores ")
    con=0
    sum=3
    for i in range(0,len(A)-1,3):
        SplineCuadratico.cell(row=sum,column=1,value=sln[i])
        SplineCuadratico.cell(row=sum,column=2,value=sln[i+1])
        SplineCuadratico.cell(row=sum,column=3,value=sln[i+2])
        SplineCuadratico.cell(row=sum,column=5,value="Intervalo ["+str(xi[con])+","+str(xi[con+1])+"]")
        #print("Intervalo ["+str(xi[con])+","+str(xi[con+1])+"]")
        #print((sln[i]*(x**2))+(sln[i+1]*x)+sln[i+2])
        con+=1
        sum+=1
    sum+=2
    SplineCuadratico.cell(row=sum,column=1,value="Trazadores")
    sum+=2
    con=0
    for i in range(0,len(A)-1,3):
        SplineCuadratico.cell(row=sum,column=1,value=str((sln[i]*(x**2))+(sln[i+1]*x)+sln[i+2]))
        SplineCuadratico.cell(row=sum,column=5,value="Intervalo ["+str(xi[con])+","+str(xi[con+1])+"]")
        sum+=1
        con+=1
    Excel.save(filename="Spline_Cuadratico.xlsx")
#splinecuadratico(xi,yi) llamar a la funcion