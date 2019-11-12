import numpy as np
from openpyxl import Workbook
Excel=Workbook()
GausSeidel=Excel.active
GausSeidel.title="Gauss_Seidel"
def gauss_seidel(A,b,x0,tol,Nmax):
    n=len(A)
    D,L,U=np.zeros((n,n)),np.zeros((n,n)),np.zeros((n,n))
    for i in range(n):
        D[i][i]=(A[i][i])
        for j in range(i+1,n):
            L[j][i]=-A[j][i]
            U[i][j]=-A[i][j]
    TG=np.linalg.inv((D-L))@U
    CG=np.linalg.inv((D-L))@b
    #print("TG\n"+str(TG))
    GausSeidel.cell(row=1,column=1,value="TG")
    for i in range(n):
        for j in range(n):
            GausSeidel.cell(row=i+2,column=j+1,value=TG[i][j])
    #print("CG\n"+str(CG))
    cont=n+3
    GausSeidel.cell(row=cont,column=1,value="CG")
    for i in range(n):
        GausSeidel.cell(row=cont+2,column=i+1,value=CG[i])
    vp,vep=np.linalg.eig(TG)
    p=max(abs(vp))
    #print("\nRadio espectral: "+str(p))
    GausSeidel.cell(row=2,column=n+2,value="Radio Espectral")
    GausSeidel.cell(row=2,column=n+3,value=p)
    cont+=4
    GausSeidel.cell(row=cont,column=1,value="Iter")
    GausSeidel.cell(row=cont,column=2,value="Error")
    for i in range(n):
        GausSeidel.cell(row=cont,column=3+i,value=("x"+str(i + 1)))
    cont+=1
    xant=x0
    for k in range(Nmax):
        xactual=(TG@xant)+CG
        #print("Iteración "+str(k+1))
        GausSeidel.cell(row=cont,column=1,value=k+1)
        SumErr=0
        for i in range(n):
            #print("x"+str(i+1)+": "+str(xactual[i]))
            GausSeidel.cell(row=cont,column=i+3,value=xactual[i])
            SumErr+=((xactual[i]-xant[i])**2)
        xant=xactual
        Error=np.sqrt(SumErr)
        #print("Error: "+str(Error))
        GausSeidel.cell(row=cont,column=2,value=Error)
        cont+=1
        if Error < tol:
            break
    Excel.save(filename="GaussSeidel.xlsx")
#gauss_seidel(A,b,x0,tol,Nmax) LLamar a la funcion