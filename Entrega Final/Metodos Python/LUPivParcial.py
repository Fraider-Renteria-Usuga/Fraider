from openpyxl import Workbook
from funciones import *
import numpy as np
Excel=Workbook()
LU_Parcial=Excel.active
LU_Parcial.title="LU_Parcial"
def LUParcial(matriz,b):
    n=len(matriz)
    P=np.identity(n)
    L=np.zeros((n, n))
    for i in range(n):
        if matriz[i][i]==0:  # Si hay un cero en la diagonal
            print("Error\nSe encontró un cero en la diagonal ")
            break
        lis=[]
        for h in range(i,n):
            lis.append(abs(matriz[h][i]))
        for maxicol in range(i,n):
            if abs(matriz[maxicol][i])==(max(lis)):
                matriz[[i, maxicol]]=matriz[[maxicol, i]]
                P[[i,maxicol]]=P[[maxicol,i]]
                L[[i,maxicol]]=L[[maxicol,i]]
        L[i][i]=1
        for j in range(i+1,n):
            mult=(matriz[j][i])/(matriz[i][i])
            L[j][i]=mult
            for k in range(n):
                matriz[j][k]=matriz[j][k]-(mult*matriz[i][k])
    U=matriz
    Bn=P@b #Multiplique la matriz con el vector
    vecz=sustProgre(L,Bn)
    sln=sustRegre(U,vecz)
    LU_Parcial.cell(row=1,column=1,value="Matriz L final ")
    for i in range(n):
        for j in range(n):
            LU_Parcial.cell(row=i+2,column=j+1,value=L[i][j])
    cont=n+3
    LU_Parcial.cell(row=cont,column=1,value="Matriz U final ")
    for i in range(n):
        for j in range(n):
            LU_Parcial.cell(row=cont+2,column=j+1,value=U[i][j])
        cont+=1
    cont+=3
    LU_Parcial.cell(row=cont,column=1,value="Matriz P final ")
    for i in range(n):
        for j in range(n):
            LU_Parcial.cell(row=cont+2,column=j+1,value=P[i][j])
        cont+=1
    cont+=3
    LU_Parcial.cell(row=cont,column=1,value="La Solución al Sistema es: ")
    xi=1
    for i in range(n):
        LU_Parcial.cell(row=cont+2,column=1,value=("x"+str(xi)))
        LU_Parcial.cell(row=cont+2,column=2,value=sln[i])
        xi+=1
        cont+=1
    Excel.save(filename="LU_Parcial.xlsx")
    return(L,U,P,sln)
#n=len(L)
#Comando para imprimir de una manera clara al usuario
#print("Matriz L\n"+str(L))
#print("Matriz U\n"+str(U))
#print("Matriz P\n"+str(P))
#print("La solución al sistema es: ")
#for i in range(n):
    #print("x"+str(i+1)+": "+str(sln[i]))