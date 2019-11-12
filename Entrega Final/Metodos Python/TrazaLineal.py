import numpy as np
from funciones import *
import sympy as sp
from openpyxl import Workbook
Excel=Workbook()
Splinelineal=Excel.active
Splinelineal.title="Spline_Lineal"
def splinelineal(xi,yi):
    x=sp.Symbol('x')
    n=2*(len(xi)-1)
    A=np.zeros((n,n))
    b=np.zeros(n)
    cont=2
    ini=0
    sucont=-1
    for i in range(0,n,2):
        for j in range(ini,cont):
            A[j][i]=xi[j]
            A[j][i+1]=1
            b[j]=yi[j]
        if ini>=2:
            ini=cont
            sucont+=1
        if ini==0:
            ini+=2
        cont+=1
    c=1
    for i in range(((len(xi))),(cont+sucont)):
        A[i]=A[c]
        c+=1
    ini2=0
    fin2=2
    for i in range(cont-1,n):
        for j in range(ini2,fin2):
            A[i][j+2]=-A[i][j]
        ini2+=2
        fin2+=2
    #print(A)
    GauSimple(A,b)
    sln=sustRegre(A,b)
    Splinelineal.cell(row=1,column=1,value="Coeficientes de los trazadores ")
    con=0
    sum=3
    for i in range(0,len(A)-1,2):
        Splinelineal.cell(row=sum,column=1,value=sln[i])
        Splinelineal.cell(row=sum,column=2,value=sln[i+1])
        Splinelineal.cell(row=sum,column=4,value="Intervalo ["+str(xi[con])+"," + str(xi[con+1])+"]")
        #print("Intervalo ["+str(xi[con])+"," + str(xi[con+1])+"]")
        #print((sln[i]*x)+sln[i+1])
        con+=1
        sum+=1
    sum+=2
    Splinelineal.cell(row=sum,column=1,value="Trazadores")
    sum+=2
    con=0
    for i in range(0,len(A)-1,2):
        Splinelineal.cell(row=sum,column=1,value=str((sln[i]*x)+sln[i+1]))
        Splinelineal.cell(row=sum,column=4,value="Intervalo ["+str(xi[con])+","+str(xi[con+1])+"]")
        sum+=1
        con+=1
    Excel.save(filename="Spline_Lineal.xlsx")
#splinelineal(xi,yi) LLamar a la funcion