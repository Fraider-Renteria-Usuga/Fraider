import numpy as np
from funciones import *
from sympy import *
from openpyxl import Workbook
Excel=Workbook()
Vandermonde=Excel.active
Vandermonde.title="Vandermonde"
def vandermonde(xi,yi):
    n=len(xi)
    x=symbols('x')
    matriz=np.zeros((n, n))
    vector=np.zeros(n)
    for i in range(n):
        cont=n-1
        for j in range(n):
            matriz[i][j]=xi[i]**(cont)
            cont-=1
            vector[i]=yi[i]
    #print(matriz)
    Vandermonde.cell(row=1,column=1,value="Matriz de vandermonde")
    for i in range(n):
        for j in range(n):
            Vandermonde.cell(row=i+2,column=j+1,value=matriz[i][j])
    matriz1,vectorSol=GauSimple(matriz,vector)
    sln=sustRegre(matriz1,vectorSol)
    cont=n+3
    Vandermonde.cell(row=cont,column=1,value="Coeficientes del polinomio ")
    cont+=2
    ex=n-1
    polinomio=0
    for i in range(n):
        Vandermonde.cell(row=cont,column=i+1,value="a"+str(n-(i+1)))
        Vandermonde.cell(row=cont+1,column=i+1,value=sln[i])
        polinomio+=(sln[i]*(x**ex))
        ex-=1
    cont+=3
    Vandermonde.cell(row=cont,column=1,value="Polinomio ")
    Vandermonde.cell(row=cont+1,column=1,value=str(polinomio))
    Excel.save(filename="Vandermonde.xlsx")
    return(matriz,sln)
#n=len(A)
#Comando para imprimir de una manera clara al usuario
#sub=n-1
#for i in sln:
    #print("a"+str(sub)+": "+str(i))
    #sub-=1
#a3x3+a2x2+a1x+a0