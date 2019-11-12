from sympy import *
import numpy as np
from openpyxl import Workbook
Excel=Workbook()
Lagrange=Excel.active
Lagrange.title="Lagrange"
def lagrange(xi,yi):
    n=len(xi)
    x=symbols('x')
    L=[]
    for i in range(n):
        num=1
        den=1
        for j in range(n):
            if xi[i]!=xi[j]:
                num*=expand(x-xi[j])
                den*=(xi[i]-xi[j])
        L.append(expand(num)/den)
    pol=0
    for i in range(n):
        pol+=yi[i]*L[i]
    Lagrange.cell(row=1,column=1,value="Polinomios interpolantes de Lagrange")
    cont=3
    for i in range(len(L)):
        Lagrange.cell(row=cont,column=1,value="L"+str(i))
        Lagrange.cell(row=cont,column=2,value=str(L[i]))
        cont+=1
    cont+=1
    Lagrange.cell(row=cont,column=1,value="Polinomio")
    Lagrange.cell(row=cont+2,column=1,value=str(pol))
    Excel.save(filename="Lagrange.xlsx")
    return(L,(expand(pol)))
#L,pol=lagrange(xi,yi) LLamar a la funcion
#Comando para imprimir de una manera clara al usuario
#for i in range(len(L)):
    #print("L" + str(i) + ": " + str(L[i]))
#print("p(x): " + str(pol))